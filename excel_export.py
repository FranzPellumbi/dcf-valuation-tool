from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Color constants
BLUE = "FF0000FF"      # inputs
BLACK = "FF000000"     # formulas
GREEN = "FF008000"     # cross-sheet links
HEADER_BG = "FF1F3864" # dark blue header
HEADER_FG = "FFFFFFFF" # white text
ALT_ROW = "FFD9E1F2"   # light blue alternating row

def style_header(cell, text):
    cell.value = text
    cell.font = Font(name="Arial", bold=True, color=HEADER_FG, size=11)
    cell.fill = PatternFill("solid", start_color=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def style_input(cell, value):
    cell.value = value
    cell.font = Font(name="Arial", color=BLUE, size=10)

def style_formula(cell, formula):
    cell.value = formula
    cell.font = Font(name="Arial", color=BLACK, size=10)

def style_label(cell, text, bold=False):
    cell.value = text
    cell.font = Font(name="Arial", bold=bold, size=10)

def thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)

def generate_excel(ticker, current_price, intrinsic_value, wacc, growth_rate,
                   terminal_growth, fcf, projected_fcf, comps_df, mc_results):
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    ws["A1"].value = f"DCF Valuation Report — {ticker}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=HEADER_FG)
    ws["A1"].fill = PatternFill("solid", start_color=HEADER_BG)
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"].value = f"Generated: {datetime.today().strftime('%B %d, %Y')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=10)

    rows = [
        ("", ""),
        ("KEY METRICS", ""),
        ("Current Price ($)", current_price),
        ("Intrinsic Value — DCF ($)", intrinsic_value),
        ("Upside / Downside", f"=(B6-B5)/B5"),
        ("", ""),
        ("DCF ASSUMPTIONS", ""),
        ("WACC", wacc),
        ("FCF Growth Rate", growth_rate),
        ("Terminal Growth Rate", terminal_growth),
        ("", ""),
        ("MONTE CARLO", ""),
        ("Median Intrinsic Value ($)", mc_results.median()),
        ("5th Percentile — Bear ($)", mc_results.quantile(0.05)),
        ("95th Percentile — Bull ($)", mc_results.quantile(0.95)),
        ("Probability Overvalued", (mc_results < current_price).mean()),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(name="Arial", bold=(value == ""), size=10)
        if isinstance(value, str) and value.startswith("="):
            style_formula(ws[f"B{i}"], value)
            ws[f"B{i}"].number_format = '0.0%'
        elif isinstance(value, float) and label in ("WACC", "FCF Growth Rate", "Terminal Growth Rate", "Probability Overvalued"):
            style_input(ws[f"B{i}"], value)
            ws[f"B{i}"].number_format = '0.0%'
        elif isinstance(value, (int, float)) and value != "":
            style_input(ws[f"B{i}"], value)
            ws[f"B{i}"].number_format = '$#,##0.00'

    # ── Sheet 2: DCF Model ────────────────────────────────────────────────
    ws2 = wb.create_sheet("DCF Model")
    ws2.column_dimensions["A"].width = 22
    for col in range(2, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    headers = ["Metric"] + [f"Year {i}" for i in range(1, 6)] + ["Terminal", "PV Total"]
    for col, h in enumerate(headers, start=1):
        style_header(ws2.cell(row=1, column=col), h)

    historical_vals = list(fcf.values)
    avg_fcf = sum(historical_vals) / len(historical_vals)

    rows_dcf = [
        ("Base FCF ($B)", [round(avg_fcf / 1e9, 2)] + [""] * 6),
        ("Growth Rate", [growth_rate] * 5 + ["", ""]),
        ("Projected FCF ($B)", [round(v / 1e9, 2) for v in projected_fcf] + ["", ""]),
        ("Discount Factor", [round(1 / (1 + wacc) ** i, 4) for i in range(1, 6)] + ["", ""]),
        ("PV of FCF ($B)", [round(v / 1e9, 2) for v in []]),
    ]

    for r, (label, values) in enumerate(rows_dcf, start=2):
        ws2[f"A{r}"] = label
        ws2[f"A{r}"].font = Font(name="Arial", size=10)
        for c, val in enumerate(values, start=2):
            cell = ws2.cell(row=r, column=c)
            style_input(cell, val)
            if "FCF" in label or "PV" in label:
                cell.number_format = '$#,##0.00'
            elif "Rate" in label or "Factor" in label:
                cell.number_format = '0.00%' if "Rate" in label else '0.0000'

    # ── Sheet 3: Historical FCF ───────────────────────────────────────────
    ws3 = wb.create_sheet("Historical FCF")
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 20

    style_header(ws3["A1"], "Date")
    style_header(ws3["B1"], "Free Cash Flow ($B)")

    for i, (date, value) in enumerate(fcf.items(), start=2):
        ws3[f"A{i}"] = str(date)[:10]
        ws3[f"A{i}"].font = Font(name="Arial", size=10)
        style_input(ws3[f"B{i}"], round(value / 1e9, 2))
        ws3[f"B{i}"].number_format = '$#,##0.00'

    # ── Sheet 4: Comps ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Comps")
    for col, width in zip(["A","B","C","D","E","F"], [10, 15, 18, 14, 12, 12]):
        ws4.column_dimensions[col].width = width

    comp_headers = ["Ticker", "Price", "Market Cap ($B)", "EV/EBITDA", "P/E", "P/FCF"]
    for col, h in enumerate(comp_headers, start=1):
        style_header(ws4.cell(row=1, column=col), h)

    for i, row in comps_df.iterrows():
        r = i + 2
        ws4[f"A{r}"] = row["Ticker"]
        ws4[f"B{r}"] = row["Price"]
        ws4[f"C{r}"] = row["Market Cap ($B)"]
        ws4[f"D{r}"] = row["EV/EBITDA"]
        ws4[f"E{r}"] = row["P/E"]
        ws4[f"F{r}"] = row["P/FCF"]
        for col in range(1, 7):
            ws4.cell(row=r, column=col).font = Font(name="Arial", size=10)
            if i % 2 == 0:
                ws4.cell(row=r, column=col).fill = PatternFill("solid", start_color=ALT_ROW)
        ws4[f"D{r}"].number_format = '0.0x'
        ws4[f"E{r}"].number_format = '0.0x'
        ws4[f"F{r}"].number_format = '0.0x'

    # ── Sheet 5: Monte Carlo ──────────────────────────────────────────────
    ws5 = wb.create_sheet("Monte Carlo")
    ws5.column_dimensions["A"].width = 25
    ws5.column_dimensions["B"].width = 20

    style_header(ws5["A1"], "Percentile")
    style_header(ws5["B1"], "Implied Value Per Share ($)")

    percentiles = [
        ("5th (Bear Case)", mc_results.quantile(0.05)),
        ("25th Percentile", mc_results.quantile(0.25)),
        ("50th — Median", mc_results.quantile(0.50)),
        ("75th Percentile", mc_results.quantile(0.75)),
        ("95th (Bull Case)", mc_results.quantile(0.95)),
        ("Mean", mc_results.mean()),
        ("Std Deviation", mc_results.std()),
    ]

    for i, (label, value) in enumerate(percentiles, start=2):
        ws5[f"A{i}"] = label
        ws5[f"A{i}"].font = Font(name="Arial", size=10)
        style_input(ws5[f"B{i}"], round(value, 2))
        ws5[f"B{i}"].number_format = '$#,##0.00'

    filename = f"{ticker}_Valuation_Model.xlsx"
    wb.save(filename)
    return filename

